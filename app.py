"""
Streamlit app: Upload NHIỀU file raw dashboard (bao nhiêu tháng cũng được),
app TỰ ĐỘNG nhận diện tháng, tự gộp và xuất ra file "Phân tích Doanh thu
khách hàng" — mỗi chi nhánh 1 sheet, mỗi tháng 1 cột, sắp xếp theo thời gian.

Điểm khác so với bản trước:
  1. Tên cột (tháng) được TỰ ĐỘNG điền và sắp xếp — không cần gõ tay.
     Chỉ khi nào app không tự đoán được ngày tháng, hoặc 2 file trùng
     nhãn tháng, mới hiện ô để bạn sửa tay (đúng chỗ bị lỗi thôi).
  2. Kiểm tra số liệu kỹ hơn — ngoài check trùng KPI và biến động
     tháng-qua-tháng, còn có thêm bước ĐỐI SOÁT TỔNG: so khớp
     "Tất cả chi nhánh" với tổng cộng dồn từng chi nhánh, để phát hiện
     copy nhầm dòng / thiếu chi nhánh trong raw dashboard.
  3. File xuất ra có thêm 1 sheet "Audit" ghi lại toàn bộ kết quả kiểm
     tra (đối soát tổng, KPI trùng, biến động lớn) để lưu vết lâu dài,
     không chỉ hiện trên UI rồi mất.

Cách chạy:
    pip install streamlit openpyxl pandas
    streamlit run app.py
"""

import io
import re
import shutil
import subprocess
import tempfile
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 1. MAPPING: cột trong RAW DASHBOARD  ->  dòng trong sheet phân tích
#    (đã đối chiếu khớp 100% với dữ liệu T6/T7/T8-2026 thực tế của bạn)
# ---------------------------------------------------------------------------

RAW_COLS = {
    "kpi": "KPI",
    "doanh_thu": "Doanh thu",
    "khach_moi": "Khách mới",
    "mua_tt_km": "Mua TT KM",
    "dt_khach_moi_all": "DT khách mới (all)",
    "bill_tb_km": "Bill TB KM",
    "dt_khach_moi_30d": "DT khách mới 30D",
    "booking_moi": "Booking Mới",
    "checkin_moi": "Checkin Mới",
    "khach_thuc_te": "Khách thực tế",   # dạng "30838 (3434 / 1361 / 26043)"
    "mua_tt_kc": "Mua TT KC",
    "dt_khach_cu": "DT khách cũ",
    "bill_tb_mua_kc": "Bill TB mua KC",
}

# (dòng trong sheet phân tích, nhãn, loại, key trong RAW_COLS)
TEMPLATE_ROWS = [
    (2, "KPI", "raw", "kpi"),
    (3, "Tổng doanh thu", "raw", "doanh_thu"),
    (4, "Khách mới", "raw", "khach_moi"),
    (5, "Khách mua hàng TT (mới)", "raw", "mua_tt_km"),
    (6, "Tỷ lệ chốt khách mới", "formula_moi", None),
    (7, "Doanh thu khách mới", "raw", "dt_khach_moi_all"),
    (8, "Bill TB khách mới", "raw", "bill_tb_km"),
    (9, "Doanh thu khách mới 30 ngày", "raw", "dt_khach_moi_30d"),
    (10, "Bill TB khách mới 30 ngày", "missing", None),
    (11, "Khách booking mới", "raw", "booking_moi"),
    (12, "Khách checkin mới", "raw", "checkin_moi"),
    (13, "Khách thực tế (cũ)", "khach_cu", None),
    (14, "Khách mua hàng TT (cũ)", "raw", "mua_tt_kc"),
    (15, "Tỷ lệ chốt khách cũ", "formula_cu", None),
    (16, "Doanh thu khách cũ", "raw", "dt_khach_cu"),
    (17, "Bill TB khách cũ", "raw", "bill_tb_mua_kc"),
]

MONEY_ROWS = {2, 3, 7, 9, 16}

# Chi nhánh không tính vào KPI kinh doanh (đào tạo / hành chính) -> không ra
# sheet riêng. "Học Viện LGS" cũng KHÔNG được cộng vào "Tất cả chi nhánh"
# trong raw dashboard, nên loại luôn khi đối soát tổng bên dưới.
EXCLUDE_BRANCHES = {"Học Viện LGS", "Văn Phòng"}
EXCLUDE_FROM_TOTAL_RECONCILIATION = {"Học Viện LGS"}

FONT = Font(name="Arial", size=11)
FONT_BOLD = Font(name="Arial", size=11, bold=True)
FONT_HEADER = Font(name="Arial", size=11, bold=True, color="FFFFFF")
FILL_HEADER = PatternFill("solid", fgColor="4472C4")
FILL_INPUT = PatternFill("solid", fgColor="FFFF00")
FILL_WARN = PatternFill("solid", fgColor="FFC7CE")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = "#,##0"
INT_FMT = "#,##0"
PCT_FMT = "0.0%"

MOM_THRESHOLD = 0.6          # ngưỡng cảnh báo biến động tháng-qua-tháng
RECONCILE_TOLERANCE = 1000   # sai số cho phép khi đối soát tổng (VNĐ)


# ---------------------------------------------------------------------------
# 2. Đọc raw dashboard + tự nhận diện tháng từ tiêu đề file
# ---------------------------------------------------------------------------

def to_num(v):
    if isinstance(v, (int, float)):
        return v
    if isinstance(v, str):
        vv = v.replace(",", "").replace("\xa0", "").strip()
        try:
            return float(vv)
        except ValueError:
            return None
    return None


def parse_khach_cu(v):
    if v is None:
        return None
    nums = re.findall(r"-?\d+", str(v))
    if len(nums) >= 4:
        return int(nums[3])
    return None


def read_raw_dashboard(file):
    """Trả về (title, {tên chi nhánh: {tên cột: giá trị}})."""
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = wb.active

    title = ws.cell(row=1, column=1).value or ""
    header_row_idx = None
    for r in range(1, 6):
        cell = ws.cell(row=r, column=1).value
        if cell and str(cell).strip() == "Chi nhánh":
            header_row_idx = r
            break
    if header_row_idx is None:
        raise ValueError("Không tìm thấy dòng tiêu đề 'Chi nhánh' trong file.")

    headers = [c.value for c in ws[header_row_idx]]
    missing_cols = [name for name in RAW_COLS.values() if name not in headers]
    if missing_cols:
        raise ValueError(
            "File raw thiếu các cột cần thiết: " + ", ".join(missing_cols) +
            ". Kiểm tra lại định dạng file export từ dashboard."
        )

    data = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        name = row[0]
        if not name:
            continue
        data[str(name).strip()] = dict(zip(headers, row))
    return title, data


def detect_month(title: str, filename: str):
    """Tìm khoảng ngày 'dd-mm-yyyy to dd-mm-yyyy' hoặc 'dd-mm-yyyy' đầu tiên
    trong tiêu đề/tên file. Trả về (year, month, ngay_bat_dau) hoặc
    (None, None, None) nếu không tìm được."""
    text = f"{title} {filename}"
    m = re.search(r"(\d{2})-(\d{2})-(\d{4})", text)
    if m:
        dd, mm, yyyy = m.groups()
        try:
            return int(yyyy), int(mm), date(int(yyyy), int(mm), int(dd))
        except ValueError:
            pass
    return None, None, None


def auto_generate_labels(parsed: list[dict]) -> list[str]:
    """Tự động sinh nhãn cột tháng cho từng file, không cần người dùng gõ tay.
    - Nếu nhận diện được tháng/năm: nhãn mặc định là 'T{thang}'.
    - Nếu 2+ file trùng nhãn (vd cùng là T1 nhưng khác năm), tự thêm hậu tố
      năm 2 số: 'T1/26', 'T1/27' để phân biệt.
    - Nếu không nhận diện được ngày tháng, dùng tên file (không dấu .xlsx)
      làm nhãn tạm — trường hợp này sẽ được đánh dấu để người dùng xác nhận.
    """
    base_labels = []
    for p in parsed:
        if p["month"]:
            base_labels.append(f"T{p['month']}")
        else:
            base_labels.append(Path(p["filename"]).stem)

    # Đếm số lần xuất hiện của từng nhãn cơ bản -> nếu >1 thì cần thêm năm
    from collections import Counter
    counts = Counter(base_labels)

    final_labels = []
    for p, base in zip(parsed, base_labels):
        if counts[base] > 1 and p["year"]:
            final_labels.append(f"{base}/{str(p['year'])[-2:]}")
        else:
            final_labels.append(base)
    return final_labels


def build_branch_values(branch_row: dict) -> dict:
    raw_vals = {k: branch_row.get(colname) for k, colname in RAW_COLS.items()}
    out = {}
    for row_idx, _label, kind, key in TEMPLATE_ROWS:
        if kind == "raw":
            out[row_idx] = to_num(raw_vals[key])
        elif kind == "khach_cu":
            out[row_idx] = parse_khach_cu(raw_vals["khach_thuc_te"])
        else:
            out[row_idx] = None
    return out


# ---------------------------------------------------------------------------
# 3. Kiểm tra bất thường (audit) — 3 lớp kiểm tra độc lập
# ---------------------------------------------------------------------------

def check_duplicate_kpi(label: str, raw_data: dict) -> list[str]:
    """Lớp 1: 2 chi nhánh có KPI giống hệt nhau -> khả năng copy nhầm dòng."""
    warnings = []
    kpi_by_branch = {
        b: to_num(v.get("KPI"))
        for b, v in raw_data.items()
        if b not in EXCLUDE_BRANCHES and b != "Tất cả chi nhánh" and to_num(v.get("KPI"))
    }
    seen = {}
    for b, kpi in kpi_by_branch.items():
        seen.setdefault(kpi, []).append(b)
    for kpi, branches in seen.items():
        if len(branches) > 1:
            warnings.append(
                f"[{label}] KPI giống hệt nhau ({kpi:,.0f}) giữa: "
                f"{', '.join(branches)} — khả năng copy nhầm dòng trong raw dashboard."
            )
    return warnings


def check_total_reconciliation(label: str, raw_data: dict) -> list[str]:
    """Lớp 2 (MỚI): đối soát dòng 'Tất cả chi nhánh' với tổng cộng dồn từng
    chi nhánh riêng lẻ. Đây là cách hiệu quả nhất để bắt lỗi thiếu chi nhánh
    hoặc copy sai số trong raw dashboard, vì HQ tổng và tổng từng chi nhánh
    phải luôn khớp nhau (trừ Học Viện LGS không tính vào tổng kinh doanh)."""
    warnings = []
    if "Tất cả chi nhánh" not in raw_data:
        warnings.append(f"[{label}] Không tìm thấy dòng 'Tất cả chi nhánh' để đối soát.")
        return warnings

    for field_key, field_name in [("doanh_thu", "Doanh thu"), ("khach_moi", "Khách mới")]:
        col = RAW_COLS[field_key]
        reported = to_num(raw_data["Tất cả chi nhánh"].get(col))
        included_sum = sum(
            to_num(v.get(col)) or 0
            for b, v in raw_data.items()
            if b not in ({"Tất cả chi nhánh"} | EXCLUDE_FROM_TOTAL_RECONCILIATION)
        )
        if reported is None:
            continue
        diff = reported - included_sum
        if abs(diff) > RECONCILE_TOLERANCE:
            warnings.append(
                f"[{label}] '{field_name}' TẤT CẢ CHI NHÁNH = {reported:,.0f} nhưng "
                f"tổng cộng dồn từng chi nhánh = {included_sum:,.0f} "
                f"(lệch {diff:,.0f}) — kiểm tra lại raw dashboard có thiếu/thừa "
                f"chi nhánh nào không."
            )
    return warnings


def check_month_over_month(sheet_title, month_labels, values_by_month) -> list[str]:
    """Lớp 3: biến động bất thường giữa các tháng liên tiếp trong CÙNG 1 chi nhánh."""
    warnings = []
    for i in range(1, len(month_labels)):
        prev_label, cur_label = month_labels[i - 1], month_labels[i]
        prev_vals, cur_vals = values_by_month[i - 1], values_by_month[i]
        for row_idx, label, kind, _key in TEMPLATE_ROWS:
            if kind not in ("raw", "khach_cu"):
                continue
            pv, cv = prev_vals.get(row_idx), cur_vals.get(row_idx)
            if pv in (None, 0) or cv is None:
                continue
            change = (cv - pv) / pv
            if abs(change) > MOM_THRESHOLD:
                warnings.append(
                    f"[{sheet_title}] '{label}': {prev_label}={pv:,.0f} → "
                    f"{cur_label}={cv:,.0f} ({change*100:+.0f}%) — biến động lớn, nên kiểm tra lại."
                )
    return warnings



# ---------------------------------------------------------------------------
# 4. Xây dựng workbook nhiều tháng (+ sheet Audit lưu vết kiểm tra)
#    File Excel xuất ra giữ nguyên như bản gốc — KHÔNG có sheet Dashboard/
#    biểu đồ trong này nữa. Toàn bộ phần trực quan (biểu đồ, thẻ KPI...)
#    giờ nằm trên chính trang web Streamlit, xem hàm render_web_dashboard().
# ---------------------------------------------------------------------------

def build_workbook(months: list[dict]):
    """months: list các dict {'label': str, 'raw_data': {...}} đã sắp xếp theo thời gian."""
    all_branches = []
    for m in months:
        for b in m["raw_data"].keys():
            if b not in EXCLUDE_BRANCHES and b not in all_branches:
                all_branches.append(b)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    all_warnings = []

    # --- Lớp 1 + 2: kiểm tra trong nội bộ từng tháng (không cần build sheet) ---
    for m in months:
        all_warnings.extend(check_duplicate_kpi(m["label"], m["raw_data"]))
        all_warnings.extend(check_total_reconciliation(m["label"], m["raw_data"]))

    for branch in all_branches:
        ws = wb.create_sheet(branch[:31])  # Excel giới hạn tên sheet 31 ký tự
        ws.column_dimensions["A"].width = 32
        ws["A1"] = "Chỉ tiêu"
        ws["A1"].font = FONT_HEADER
        ws["A1"].fill = FILL_HEADER
        ws["A1"].border = BORDER

        values_by_month = []
        for col_idx, m in enumerate(months, start=2):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 18
            header_cell = ws[f"{col_letter}1"]
            header_cell.value = m["label"]
            header_cell.font = FONT_HEADER
            header_cell.fill = FILL_HEADER
            header_cell.alignment = Alignment(horizontal="center")
            header_cell.border = BORDER

            branch_row = m["raw_data"].get(branch, {})
            computed = build_branch_values(branch_row) if branch_row else {}
            values_by_month.append(computed)

            for row_idx, label, kind, _key in TEMPLATE_ROWS:
                a = ws[f"A{row_idx}"]
                a.value = label
                a.font = FONT_BOLD if kind.startswith("formula") else FONT
                a.border = BORDER

                cell = ws[f"{col_letter}{row_idx}"]
                cell.font = FONT
                cell.border = BORDER
                cell.alignment = Alignment(horizontal="right")

                if kind == "formula_moi":
                    cell.value = f"={col_letter}5/{col_letter}4"
                    cell.number_format = PCT_FMT
                elif kind == "formula_cu":
                    cell.value = f"={col_letter}14/{col_letter}13"
                    cell.number_format = PCT_FMT
                elif kind == "missing":
                    cell.value = None
                    cell.fill = FILL_INPUT
                    cell.comment = Comment(
                        "Raw dashboard không có số liệu này. Vui lòng nhập tay.",
                        "App phân tích doanh thu",
                    )
                else:
                    val = computed.get(row_idx)
                    cell.value = val
                    cell.number_format = MONEY_FMT if row_idx in MONEY_ROWS else INT_FMT
                    if not branch_row:
                        cell.fill = FILL_INPUT
                        cell.comment = Comment(
                            f"Không có dữ liệu chi nhánh này trong file raw tháng {m['label']}.",
                            "App phân tích doanh thu",
                        )

        ws.freeze_panes = "B2"
        month_labels = [m["label"] for m in months]
        all_warnings.extend(
            check_month_over_month(branch, month_labels, values_by_month)
        )

    # --- Sheet Audit: lưu toàn bộ kết quả kiểm tra vào chính file xuất ra ---
    audit_ws = wb.create_sheet("Audit", 0)
    audit_ws.column_dimensions["A"].width = 100
    audit_ws["A1"] = f"Báo cáo kiểm tra số liệu — tự động tạo lúc xuất file"
    audit_ws["A1"].font = FONT_BOLD
    audit_ws["A2"] = f"Các tháng trong file: {', '.join(m['label'] for m in months)}"
    audit_ws["A2"].font = FONT
    r = 4
    if all_warnings:
        audit_ws[f"A{r}"] = f"⚠ Phát hiện {len(all_warnings)} điểm cần kiểm tra lại:"
        audit_ws[f"A{r}"].font = FONT_BOLD
        r += 1
        for w in all_warnings:
            cell = audit_ws[f"A{r}"]
            cell.value = w
            cell.font = FONT
            cell.fill = FILL_WARN
            cell.alignment = Alignment(wrap_text=True)
            r += 1
    else:
        audit_ws[f"A{r}"] = "✓ Không phát hiện bất thường nào ở tất cả các bước kiểm tra."
        audit_ws[f"A{r}"].font = FONT
        r += 1
    audit_ws.freeze_panes = "A4"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, all_warnings


def recalc_with_libreoffice(xlsx_bytes: io.BytesIO, timeout: int = 30) -> io.BytesIO:
    """Mở/lưu lại file bằng LibreOffice headless để công thức (=B5/B4...) có
    sẵn giá trị đã tính (cached value), thay vì để trống cho tới khi người
    dùng tự mở file và tính lại. Nếu máy chủ không có LibreOffice, hoặc có
    lỗi bất kỳ, trả nguyên file gốc (công thức vẫn đúng, Excel sẽ tự tính khi
    mở bình thường — chỉ là chưa có sẵn giá trị cache)."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return xlsx_bytes

    try:
        with tempfile.TemporaryDirectory() as tmpdir:
            in_dir = Path(tmpdir) / "in"
            out_dir = Path(tmpdir) / "out"
            in_dir.mkdir()
            out_dir.mkdir()
            in_path = in_dir / "in.xlsx"
            in_path.write_bytes(xlsx_bytes.getvalue())
            # outdir PHẢI khác thư mục chứa file gốc, nếu không LibreOffice
            # sẽ báo lỗi ghi đè (SfxBaseModel) và không recalc được.
            result = subprocess.run(
                [soffice, "--headless", "--calc", "--convert-to", "xlsx",
                 "--outdir", str(out_dir), str(in_path)],
                capture_output=True, timeout=timeout, text=True,
            )
            out_path = out_dir / "in.xlsx"
            if result.returncode == 0 and out_path.exists():
                recalced = io.BytesIO(out_path.read_bytes())
                recalced.seek(0)
                return recalced
    except Exception:
        pass
    return xlsx_bytes


# ---------------------------------------------------------------------------
# 5. Dashboard TRỰC TIẾP TRÊN WEB (Streamlit + Plotly) — không đụng gì tới
#    file Excel xuất ra. Toàn bộ phần này chỉ hiển thị trên trình duyệt.
# ---------------------------------------------------------------------------

STAT_FIELDS = [
    ("kpi", "KPI", 2, MONEY_FMT),
    ("doanh_thu", "Doanh thu", 3, MONEY_FMT),
    ("khach_moi", "Khách mới", 4, INT_FMT),
    ("mua_tt_km", "Mua TT (mới)", 5, INT_FMT),
    ("dt_khach_moi", "DT khách mới", 7, MONEY_FMT),
    ("bill_tb_km", "Bill TB (mới)", 8, MONEY_FMT),
    ("dt_khach_moi_30d", "DT khách mới 30 ngày", 9, MONEY_FMT),
    ("booking_moi", "Booking mới", 11, INT_FMT),
    ("checkin_moi", "Checkin mới", 12, INT_FMT),
    ("khach_cu", "Khách cũ (thực tế)", 13, INT_FMT),
    ("mua_tt_kc", "Mua TT (cũ)", 14, INT_FMT),
    ("dt_khach_cu", "DT khách cũ", 16, MONEY_FMT),
    ("bill_tb_kc", "Bill TB (cũ)", 17, MONEY_FMT),
]
STAT_ROW_BY_KEY = {key: row for key, _label, row, _fmt in STAT_FIELDS}


def build_stats_dataframe(months: list[dict], all_branches: list[str]) -> pd.DataFrame:
    """Gộp toàn bộ số liệu (mọi tháng, mọi chi nhánh) thành 1 bảng 'dài'
    (tidy dataframe) để vẽ biểu đồ Plotly và làm bảng thống kê trên web.
    Đây KHÔNG phải dữ liệu ghi vào Excel — chỉ dùng nội bộ để hiển thị."""
    records = []
    for m in months:
        for branch in all_branches:
            branch_row = m["raw_data"].get(branch)
            if not branch_row:
                continue
            computed = build_branch_values(branch_row)
            rec = {"Tháng": m["label"], "Chi nhánh": branch}
            for key, _label, row_idx, _fmt in STAT_FIELDS:
                rec[key] = computed.get(row_idx)
            rec["ty_le_chot_moi"] = (
                rec["mua_tt_km"] / rec["khach_moi"] if rec.get("khach_moi") else None
            )
            rec["ty_le_chot_cu"] = (
                rec["mua_tt_kc"] / rec["khach_cu"] if rec.get("khach_cu") else None
            )
            records.append(rec)
    df = pd.DataFrame(records)
    if not df.empty:
        # giữ đúng thứ tự tháng theo months (không để pandas tự sắp xếp theo chữ cái)
        month_order = [m["label"] for m in months]
        df["Tháng"] = pd.Categorical(df["Tháng"], categories=month_order, ordered=True)
    return df


def _hex_to_rgb(h: str):
    h = h.lstrip("#")
    return tuple(int(h[i:i + 2], 16) for i in (0, 2, 4))


def _lerp_color(c1, c2, t):
    return tuple(int(c1[i] + (c2[i] - c1[i]) * t) for i in range(3))


def _normalize(series: pd.Series):
    s = series.astype(float)
    lo, hi = s.min(), s.max()
    if pd.isna(lo) or pd.isna(hi) or hi == lo:
        return [0.5] * len(s)
    return ((s - lo) / (hi - lo)).tolist()


def _green_gradient(series: pd.Series):
    """Tô nền màu xanh lá đậm dần theo giá trị — thay thế cho
    Styler.background_gradient (cần matplotlib, Streamlit Cloud không có
    sẵn) bằng cách tự nội suy màu, không phụ thuộc thư viện ngoài."""
    light, dark = _hex_to_rgb("EDF8E9"), _hex_to_rgb("006D2C")
    return [
        f"background-color: rgb{_lerp_color(light, dark, t)}; color: {'white' if t > 0.6 else 'black'}"
        for t in _normalize(series)
    ]


def _red_yellow_green_gradient(series: pd.Series):
    """Tô nền đỏ (thấp) → vàng (giữa) → xanh lá (cao), không cần matplotlib."""
    red, yellow, green = _hex_to_rgb("F8696B"), _hex_to_rgb("FFEB84"), _hex_to_rgb("63BE7B")
    out = []
    for t in _normalize(series):
        if t < 0.5:
            color = _lerp_color(red, yellow, t / 0.5)
        else:
            color = _lerp_color(yellow, green, (t - 0.5) / 0.5)
        out.append(f"background-color: rgb{color}")
    return out


def render_web_dashboard(df: pd.DataFrame, months: list[dict]):
    """Vẽ Dashboard đầy đủ ngay trên trang web bằng Plotly — thay thế hoàn
    toàn cho việc phải mở Excel mới xem được biểu đồ."""
    if df.empty:
        st.info("Chưa có đủ dữ liệu để vẽ Dashboard.")
        return

    latest_label = months[-1]["label"]
    prev_label = months[-2]["label"] if len(months) >= 2 else None
    has_total = (df["Chi nhánh"] == "Tất cả chi nhánh").any()
    df_branches = df[df["Chi nhánh"] != "Tất cả chi nhánh"].copy()
    df_total = df[df["Chi nhánh"] == "Tất cả chi nhánh"].copy()

    st.header("📊 Dashboard trực quan")
    st.caption(
        f"Các tháng: {', '.join(m['label'] for m in months)} • "
        "Biểu đồ dưới đây chỉ hiển thị trên web — không ảnh hưởng tới file Excel xuất ra."
    )

    # === HÀNG THẺ KPI ===
    if has_total:
        latest_total = df_total[df_total["Tháng"] == latest_label].iloc[0]
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Tổng doanh thu " + f"({latest_label})", f"{latest_total['doanh_thu']:,.0f} đ")
        pct_kpi = latest_total["doanh_thu"] / latest_total["kpi"] if latest_total["kpi"] else None
        c2.metric("% Hoàn thành KPI", f"{pct_kpi*100:,.1f}%" if pct_kpi is not None else "n/a")
        c3.metric("Khách mới", f"{latest_total['khach_moi']:,.0f}")
        if prev_label:
            prev_total = df_total[df_total["Tháng"] == prev_label]
            if not prev_total.empty and prev_total.iloc[0]["doanh_thu"]:
                growth = (latest_total["doanh_thu"] - prev_total.iloc[0]["doanh_thu"]) / prev_total.iloc[0]["doanh_thu"]
                c4.metric(f"Tăng trưởng DT so với {prev_label}", f"{growth*100:+.1f}%")
            else:
                c4.metric("Tăng trưởng DT", "n/a")
        else:
            c4.metric("Tăng trưởng DT", "n/a")

    tabs = st.tabs([
        "🏢 So sánh Chi nhánh", "📈 Xu hướng", "🧩 Cơ cấu & Phễu",
        "🎯 Tỷ lệ chốt & Bill TB", "📋 Bảng chi tiết",
    ])

    # --- TAB 1: So sánh chi nhánh ---
    with tabs[0]:
        fig1 = px.bar(
            df_branches, x="Chi nhánh", y="doanh_thu", color="Tháng", barmode="group",
            title="Doanh thu theo Chi nhánh qua các tháng", labels={"doanh_thu": "Doanh thu (đ)"},
        )
        fig1.update_layout(yaxis_tickformat=",.0f")
        st.plotly_chart(fig1, use_container_width=True)

        df_latest_branches = df_branches[df_branches["Tháng"] == latest_label].copy()
        df_latest_branches["pct_kpi"] = df_latest_branches["doanh_thu"] / df_latest_branches["kpi"]
        df_latest_branches = df_latest_branches.sort_values("pct_kpi", ascending=True)
        fig2 = px.bar(
            df_latest_branches, x="pct_kpi", y="Chi nhánh", orientation="h",
            color="pct_kpi", color_continuous_scale="RdYlGn",
            title=f"% Hoàn thành KPI theo Chi nhánh ({latest_label})",
            labels={"pct_kpi": "% Hoàn thành KPI"},
        )
        fig2.update_layout(xaxis_tickformat=".0%", coloraxis_showscale=False)
        st.plotly_chart(fig2, use_container_width=True)

    # --- TAB 2: Xu hướng ---
    with tabs[1]:
        if has_total:
            fig3 = go.Figure()
            fig3.add_trace(go.Scatter(x=df_total["Tháng"], y=df_total["kpi"], name="KPI", mode="lines+markers"))
            fig3.add_trace(go.Scatter(x=df_total["Tháng"], y=df_total["doanh_thu"], name="Doanh thu thực tế", mode="lines+markers"))
            fig3.update_layout(title="Doanh thu thực tế vs KPI theo tháng", yaxis_tickformat=",.0f")
            st.plotly_chart(fig3, use_container_width=True)

            fig4 = go.Figure()
            fig4.add_trace(go.Scatter(x=df_total["Tháng"], y=df_total["khach_moi"], name="Khách mới", mode="lines+markers"))
            fig4.add_trace(go.Scatter(x=df_total["Tháng"], y=df_total["mua_tt_kc"], name="Khách cũ mua hàng", mode="lines+markers"))
            fig4.update_layout(title="Xu hướng Khách mới vs Khách cũ mua hàng")
            st.plotly_chart(fig4, use_container_width=True)

            fig5 = px.area(
                df_total, x="Tháng", y="dt_khach_moi_30d",
                title="Doanh thu Khách mới trong 30 ngày đầu theo tháng",
                labels={"dt_khach_moi_30d": "Doanh thu (đ)"},
            )
            fig5.update_layout(yaxis_tickformat=",.0f")
            st.plotly_chart(fig5, use_container_width=True)
        else:
            st.info("Không tìm thấy dòng 'Tất cả chi nhánh' để vẽ biểu đồ xu hướng toàn công ty.")

    # --- TAB 3: Cơ cấu & Phễu ---
    with tabs[2]:
        if has_total:
            fig6 = px.bar(
                df_total, x="Tháng", y=["dt_khach_moi", "dt_khach_cu"],
                title="Cơ cấu Doanh thu: Khách mới vs Khách cũ theo tháng",
                labels={"value": "Doanh thu (đ)", "variable": "Loại khách"},
            )
            fig6.update_layout(yaxis_tickformat=",.0f")
            st.plotly_chart(fig6, use_container_width=True)

            col_a, col_b = st.columns(2)
            with col_a:
                latest_total_row = df_total[df_total["Tháng"] == latest_label].iloc[0]
                fig7 = px.pie(
                    names=["Khách mới", "Khách cũ"],
                    values=[latest_total_row["dt_khach_moi"], latest_total_row["dt_khach_cu"]],
                    title=f"Cơ cấu DT Khách mới/cũ — {latest_label}",
                )
                st.plotly_chart(fig7, use_container_width=True)
            with col_b:
                latest_total_row = df_total[df_total["Tháng"] == latest_label].iloc[0]
                fig8 = go.Figure(go.Funnel(
                    y=["Booking mới", "Checkin mới", "Mua hàng TT (mới)"],
                    x=[latest_total_row["booking_moi"], latest_total_row["checkin_moi"], latest_total_row["mua_tt_km"]],
                ))
                fig8.update_layout(title=f"Phễu chuyển đổi Khách mới ({latest_label})")
                st.plotly_chart(fig8, use_container_width=True)
        else:
            st.info("Không tìm thấy dòng 'Tất cả chi nhánh' để vẽ cơ cấu/phễu chuyển đổi.")

    # --- TAB 4: Tỷ lệ chốt & Bill trung bình ---
    with tabs[3]:
        df_latest_branches = df_branches[df_branches["Tháng"] == latest_label]
        fig9 = px.bar(
            df_latest_branches, x="Chi nhánh", y=["ty_le_chot_moi", "ty_le_chot_cu"], barmode="group",
            title=f"Tỷ lệ chốt Khách mới vs Khách cũ theo Chi nhánh ({latest_label})",
            labels={"value": "Tỷ lệ chốt", "variable": "Loại khách"},
        )
        fig9.update_layout(yaxis_tickformat=".0%")
        st.plotly_chart(fig9, use_container_width=True)

        fig10 = px.bar(
            df_latest_branches, x="Chi nhánh", y=["bill_tb_km", "bill_tb_kc"], barmode="group",
            title=f"Bill trung bình: Khách mới vs Khách cũ theo Chi nhánh ({latest_label})",
            labels={"value": "Bill trung bình (đ)", "variable": "Loại khách"},
        )
        fig10.update_layout(yaxis_tickformat=",.0f")
        st.plotly_chart(fig10, use_container_width=True)

    # --- TAB 5: Bảng chi tiết + bảng xếp hạng ---
    with tabs[4]:
        st.subheader(f"Bảng tổng hợp đầy đủ chỉ số theo Chi nhánh — {latest_label}")
        df_table = df_branches[df_branches["Tháng"] == latest_label].copy()
        display_cols = ["Chi nhánh"] + [key for key, _l, _r, _f in STAT_FIELDS]
        rename_map = {key: label for key, label, _r, _f in STAT_FIELDS}
        df_table = df_table[display_cols].rename(columns=rename_map).set_index("Chi nhánh")
        money_cols = [label for key, label, _r, fmt in STAT_FIELDS if fmt == MONEY_FMT]
        int_cols = [label for key, label, _r, fmt in STAT_FIELDS if fmt == INT_FMT]
        styled = df_table.style.format({c: "{:,.0f}" for c in money_cols + int_cols}).apply(
            _green_gradient, subset=["Doanh thu"]
        )
        st.dataframe(styled, use_container_width=True)

        st.subheader("Bảng xếp hạng nhanh theo Chi nhánh")
        rank_df = df_branches[df_branches["Tháng"] == latest_label][["Chi nhánh", "doanh_thu", "kpi"]].copy()
        rank_df["% Hoàn thành KPI"] = rank_df["doanh_thu"] / rank_df["kpi"]
        if prev_label:
            prev_rev = df_branches[df_branches["Tháng"] == prev_label][["Chi nhánh", "doanh_thu"]].rename(
                columns={"doanh_thu": "doanh_thu_prev"}
            )
            rank_df = rank_df.merge(prev_rev, on="Chi nhánh", how="left")
            rank_df["Tăng trưởng DT"] = (rank_df["doanh_thu"] - rank_df["doanh_thu_prev"]) / rank_df["doanh_thu_prev"]
            rank_df = rank_df.drop(columns=["doanh_thu_prev"])
        rank_df = rank_df.rename(columns={"doanh_thu": f"Doanh thu ({latest_label})"}).drop(columns=["kpi"])
        rank_df = rank_df.sort_values(f"Doanh thu ({latest_label})", ascending=False).set_index("Chi nhánh")
        fmt_dict = {f"Doanh thu ({latest_label})": "{:,.0f}", "% Hoàn thành KPI": "{:.1%}"}
        if "Tăng trưởng DT" in rank_df.columns:
            fmt_dict["Tăng trưởng DT"] = "{:+.1%}"
        styled_rank = rank_df.style.format(fmt_dict).apply(
            _red_yellow_green_gradient, subset=["% Hoàn thành KPI"]
        )
        st.dataframe(styled_rank, use_container_width=True)


# ---------------------------------------------------------------------------
# 6. UI
# ---------------------------------------------------------------------------

st.set_page_config(page_title="Phân tích Doanh thu khách hàng", layout="wide")
st.title("📊 Phân tích Doanh thu khách hàng — nhiều tháng")

st.markdown(
    """
Upload **1 hoặc nhiều file raw dashboard** (mỗi file là 1 tháng) — app tự
nhận diện tháng, tự đặt tên cột, gộp lại và xuất ra file **Phân tích Doanh
thu khách hàng** đầy đủ (mỗi chi nhánh 1 sheet, mỗi tháng 1 cột, xếp theo
thời gian), kèm 1 sheet **Audit** ghi lại toàn bộ kết quả kiểm tra số liệu.
Lần sau có thêm tháng mới, chỉ cần upload thêm — không giới hạn số tháng.
"""
)

raw_files = st.file_uploader(
    "Upload raw dashboard (chọn nhiều file cùng lúc được)",
    type=["xlsx"],
    accept_multiple_files=True,
)

if raw_files:
    parsed = []
    for f in raw_files:
        try:
            title, raw_data = read_raw_dashboard(f)
        except Exception as e:
            st.error(f"Lỗi đọc file '{f.name}': {e}")
            continue
        year, month, start_date = detect_month(title, f.name)
        parsed.append(
            {
                "filename": f.name,
                "title": title,
                "raw_data": raw_data,
                "year": year,
                "month": month,
                "start_date": start_date,
            }
        )

    if parsed:
        # sắp theo ngày bắt đầu nếu nhận diện được, file không nhận diện được xếp cuối
        parsed.sort(key=lambda p: (p["start_date"] is None, p["start_date"]))

        # === TỰ ĐỘNG sinh nhãn cột — không cần người dùng gõ tay ===
        labels = auto_generate_labels(parsed)
        undetected = [p["filename"] for p in parsed if p["month"] is None]

        st.subheader("✅ Đã tự động nhận diện tháng")
        preview_cols = st.columns([3, 2, 2])
        preview_cols[0].markdown("**File**")
        preview_cols[1].markdown("**Tiêu đề trong file**")
        preview_cols[2].markdown("**Tên cột (tự động)**")
        for p, label in zip(parsed, labels):
            c = st.columns([3, 2, 2])
            c[0].write(f"📄 {p['filename']}")
            c[1].write(p["title"][:45] + ("..." if len(p["title"]) > 45 else ""))
            c[2].write(f"`{label}`")

        # Chỉ hiện ô sửa tay khi thực sự cần: nhãn trùng nhau hoặc không
        # nhận diện được ngày tháng — mọi trường hợp bình thường đều tự động.
        needs_manual_fix = undetected or (len(set(labels)) != len(labels))
        if needs_manual_fix:
            st.warning(
                "⚠️ Một số file app không tự đặt tên cột chắc chắn được — "
                "vui lòng xác nhận/sửa lại tên cột bên dưới."
            )
            fixed_labels = []
            for i, (p, label) in enumerate(zip(parsed, labels)):
                new_label = st.text_input(
                    f"Tên cột cho {p['filename']}", value=label, key=f"label_fix_{i}"
                )
                fixed_labels.append(new_label)
            labels = fixed_labels
            if len(set(labels)) != len(labels):
                st.error("❌ Vẫn còn 2 file trùng tên cột — vui lòng sửa lại cho khác nhau.")
                st.stop()

        months = [
            {"label": labels[i], "raw_data": p["raw_data"]}
            for i, p in enumerate(parsed)
        ]

        # === DASHBOARD TRỰC TIẾP TRÊN WEB (không ảnh hưởng file Excel xuất ra) ===
        all_branches_preview = []
        for m in months:
            for b in m["raw_data"].keys():
                if b not in EXCLUDE_BRANCHES and b not in all_branches_preview:
                    all_branches_preview.append(b)
        stats_df = build_stats_dataframe(months, all_branches_preview)
        st.divider()
        render_web_dashboard(stats_df, months)

        st.divider()
        st.subheader("🔎 Kiểm tra số liệu (chạy tự động)")

        prelim_warnings = []
        for m in months:
            prelim_warnings.extend(check_duplicate_kpi(m["label"], m["raw_data"]))
            prelim_warnings.extend(check_total_reconciliation(m["label"], m["raw_data"]))

        if prelim_warnings:
            for w in prelim_warnings:
                st.warning(w)
        else:
            st.success(
                "Không phát hiện bất thường: không trùng KPI giữa các chi nhánh, "
                "và 'Tất cả chi nhánh' khớp với tổng cộng dồn từng chi nhánh."
            )

        st.info(
            "ℹ️ Dòng **'Bill TB khách mới 30 ngày'** không có trong raw dashboard "
            "nên sẽ để trống + tô vàng để bạn nhập tay và kiểm tra kỹ."
        )

        st.divider()
        if st.button("🚀 Xuất file phân tích", type="primary"):
            out_bytes, mom_warnings = build_workbook(months)
            out_bytes = recalc_with_libreoffice(out_bytes)
            new_only = [w for w in mom_warnings if w not in prelim_warnings]
            if new_only:
                st.subheader("⚠️ Biến động lớn giữa các tháng liên tiếp")
                for w in new_only:
                    st.warning(w)
            st.success(
                "Đã tạo file thành công! Toàn bộ cảnh báo (nếu có) cũng đã được "
                "lưu vào sheet 'Audit' trong file để tiện tra cứu sau này."
            )
            st.download_button(
                "⬇️ Tải file Phân tích Doanh thu khách hàng",
                data=out_bytes,
                file_name="Phan_tich_Doanh_thu_khach_hang.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
else:
    st.info("Vui lòng upload ít nhất 1 file raw dashboard.")
